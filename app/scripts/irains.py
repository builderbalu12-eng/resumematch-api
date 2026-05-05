#!/usr/bin/env python3
"""
IRAINS — All 10 Cumulative APIs → Excel (first 2 rows each, own sheet per API)
pip install requests openpyxl pandas
python irains_cumulative_preview.py
"""

import requests, urllib3
from datetime import date
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

BASE_URL   = "https://irainshydro.imd.gov.in:3000"
TODAY      = str(date.today())
START_DATE = "2026-04-01"
END_DATE   = TODAY
OUTPUT     = f"IRAINS_Cumulative_Preview_{TODAY}.xlsx"

ENDPOINTS = [
    ("UP AWS",          "/api/v1/up-aws/cumulative"),
    ("NHP AWS",         "/api/v1/nhp-aws/cumulative"),
    ("Zomato AWS",      "/api/v1/zomato-aws/cumulative"),
    ("Meghalaya AWS",   "/api/v1/meghalaya-aws/cumulative"),
    ("Mizoram AWS",     "/api/v1/mizoram-aws/cumulative"),
    ("Tamilnadu AWS",   "/api/v1/tamilnadu-aws/cumulative"),
    ("Uttarakhand AWS", "/api/v1/uttarakhand-aws/cumulative"),
    ("Telangana AWS",   "/api/v1/telangana-aws/cumulative"),
    ("Karnataka AWS",   "/api/v1/karnataka-aws/cumulative"),
    ("IITM Mumbai ARG", "/api/v1/iitm-mumbai/cumulative"),
]

BODY = {"startDate": START_DATE, "endDate": END_DATE}

HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
ALT_FILL   = PatternFill("solid", fgColor="D9E8F5")
ERR_FILL   = PatternFill("solid", fgColor="FFE0E0")
HDR_FONT   = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=13, color="1F4E79")
INFO_FONT  = Font(italic=True, color="888888", size=9)
THIN       = Side(style="thin", color="AAAAAA")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fetch(path):
    try:
        r = requests.post(BASE_URL + path, json=BODY, timeout=20, verify=False)
        r.raise_for_status()
        return r.json().get("data", []), None
    except Exception as e:
        return [], str(e)


def write_sheet(ws, df, network, path, total_records, error=None):
    # ── Row 1: Title ──
    ws.row_dimensions[1].height = 24
    t = ws.cell(1, 1, f"IRAINS  |  {network}  |  Cumulative Rainfall  |  {START_DATE} → {END_DATE}")
    t.font = TITLE_FONT
    t.alignment = Alignment(horizontal="left", vertical="center")

    # ── Row 2: Meta info ──
    ws.row_dimensions[2].height = 16
    meta = f"Endpoint: POST {BASE_URL}{path}   |   Body: {BODY}   |   Total records in API: {total_records}   |   Showing: first 2 rows   |   Generated: {TODAY}"
    m = ws.cell(2, 1, meta)
    m.font = INFO_FONT

    # ── Row 3: blank spacer ──
    ws.row_dimensions[3].height = 6

    # ── Error state ──
    if error:
        e = ws.cell(4, 1, f"⚠  ERROR calling API: {error}")
        e.fill = ERR_FILL
        e.font = Font(bold=True, color="CC0000")
        return

    # ── Empty state ──
    if df.empty:
        e = ws.cell(4, 1, "⚠  No data returned for this date range. Try a different startDate.")
        e.font = Font(italic=True, color="AA6600")
        return

    # ── Row 4: Column headers ──
    ws.row_dimensions[4].height = 20
    for col_num, col_name in enumerate(df.columns, 1):
        cell = ws.cell(4, col_num, col_name)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = BORDER

    # ── Rows 5–6: First 2 data rows ──
    for r_idx, (_, row) in enumerate(df.iterrows()):
        fill = ALT_FILL if r_idx % 2 == 0 else PatternFill()
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(5 + r_idx, col_num, value)
            cell.fill      = fill
            cell.border    = BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.font      = Font(size=10)
        ws.row_dimensions[5 + r_idx].height = 18

    # ── Auto column width ──
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max(
            (len(str(cell.value or "")) for cell in col_cells),
            default=10
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)

    # ── Freeze below header ──
    ws.freeze_panes = "A5"


def main():
    print(f"\n{'='*60}")
    print(f"  IRAINS — Cumulative Preview (first 2 rows per API)")
    print(f"  Range : {START_DATE} → {END_DATE}")
    print(f"  Output: {OUTPUT}")
    print(f"{'='*60}\n")

    summary_rows = []

    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:

        for network, path in ENDPOINTS:
            print(f"  Fetching  {network:20s} ...", end=" ", flush=True)
            data, err = fetch(path)

            total_records = len(data)

            if err:
                print(f"✗  ERROR")
                df_preview = pd.DataFrame()
            elif not data:
                print(f"⚠  EMPTY")
                df_preview = pd.DataFrame()
            else:
                df_preview = pd.DataFrame(data[:2])   # ← ONLY FIRST 2 ROWS
                print(f"✓  {total_records} total  →  showing 2  |  cols: {list(df_preview.columns)}")

            summary_rows.append({
                "Network":        network,
                "Endpoint":       f"{BASE_URL}{path}",
                "Total Records":  total_records,
                "Columns":        ", ".join(df_preview.columns.tolist()) if not df_preview.empty else "—",
                "Status":         "ERROR" if err else ("EMPTY" if df_preview.empty else "OK"),
                "Error":          err or "",
            })

            sheet_name = network[:31]

            # Write the 2-row preview starting at row 4 (rows 1-3 used by title/meta)
            if not df_preview.empty:
                df_preview.to_excel(writer, sheet_name=sheet_name, index=False, startrow=3)
            else:
                pd.DataFrame().to_excel(writer, sheet_name=sheet_name, index=False, startrow=3)

            write_sheet(writer.sheets[sheet_name], df_preview, network, path, total_records, error=err)

        # ── SUMMARY sheet ──
        summary_df = pd.DataFrame(summary_rows)
        summary_df.to_excel(writer, sheet_name="SUMMARY", index=False, startrow=3)
        write_sheet(writer.sheets["SUMMARY"], summary_df, "All Networks — Summary", "/summary", len(summary_df))

    print(f"\n✅  Done → {OUTPUT}")
    print(f"    {len(ENDPOINTS)} sheets + 1 SUMMARY sheet\n")


if __name__ == "__main__":
    main()